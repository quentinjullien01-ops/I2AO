"""Export XLSX d'une DPGF (BPU rempli + DQE chiffré).

Sortie : un classeur Excel avec deux feuilles :
  - "BPU" : catalogue complet, prix unitaires HT, structuré par catégorie
  - "DQE" : programme indicatif × prix → totaux par catégorie + total général
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .dpgf_engine import DPGFGeneree


COULEUR_HEADER = "1A3D6E"
COULEUR_CATEGORIE = "E8EDF5"
COULEUR_TOTAL_CAT = "F4F6FA"
COULEUR_TOTAL_GEN = "1A3D6E"

THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def exporter_dpgf_xlsx(
    dpgf: DPGFGeneree, output_path: Path, marche_ref: str = "", candidat: str = ""
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    _sheet_recap(wb, dpgf, marche_ref, candidat)
    _sheet_dqe(wb, dpgf, marche_ref, candidat)
    _sheet_bpu(wb, dpgf, marche_ref, candidat)

    wb.save(str(output_path))
    return output_path


def _sheet_recap(wb: Workbook, dpgf: DPGFGeneree, marche_ref: str, candidat: str) -> None:
    """Feuille de récapitulatif placée en première position."""
    ws = wb.create_sheet("Récapitulatif")

    for col, w in zip(("A", "B", "C"), (4, 50, 22)):
        ws.column_dimensions[col].width = w

    # Titre
    ws["A1"] = "DPGF — Récapitulatif"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=COULEUR_HEADER)
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 28

    row = 2
    if marche_ref:
        ws.cell(row=row, column=1, value=f"Marché : {marche_ref}")
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, italic=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
    if candidat:
        ws.cell(row=row, column=1, value=f"Candidat : {candidat}")
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, italic=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
    ws.cell(row=row, column=1, value=f"Date : {date.today().strftime('%d/%m/%Y')}")
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # Programme indicatif
    if dpgf.description_programme:
        ws.cell(row=row, column=1, value="Programme indicatif retenu pour le DQE :")
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True, color=COULEUR_HEADER)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        ws.cell(row=row, column=1, value=dpgf.description_programme)
        cell = ws.cell(row=row, column=1)
        cell.font = Font(name="Calibri", size=10, italic=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 36
        row += 2

    # Sous-totaux par catégorie
    ws.cell(row=row, column=1, value="").value = None
    ws.cell(row=row, column=2, value="Sous-totaux par chapitre")
    ws.cell(row=row, column=2).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=COULEUR_HEADER)
    ws.cell(row=row, column=3, value="Montant HT")
    ws.cell(row=row, column=3).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=COULEUR_HEADER)
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")
    ws.row_dimensions[row].height = 22
    for col in (1, 2, 3):
        ws.cell(row=row, column=col).border = BORDER
    row += 1

    sous_totaux: dict[str, float] = defaultdict(float)
    for ligne in dpgf.dqe:
        sous_totaux[ligne.categorie] += ligne.montant

    alt = False
    for cat, mt in sous_totaux.items():
        ws.cell(row=row, column=1, value="").value = None
        ws.cell(row=row, column=2, value=cat).font = Font(name="Calibri", size=10)
        cell_mt = ws.cell(row=row, column=3, value=round(mt, 2))
        cell_mt.number_format = "#,##0.00 €"
        cell_mt.alignment = Alignment(horizontal="right")
        cell_mt.font = Font(name="Calibri", size=10)
        bg = COULEUR_TOTAL_CAT if alt else "FFFFFF"
        for col in (1, 2, 3):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=bg)
            ws.cell(row=row, column=col).border = BORDER
        alt = not alt
        row += 1

    # Total général
    ws.cell(row=row, column=2, value="TOTAL DQE HT")
    ws.cell(row=row, column=2).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=COULEUR_TOTAL_GEN)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right", vertical="center")
    cell_tot = ws.cell(row=row, column=3, value=dpgf.montant_dqe_he)
    cell_tot.number_format = "#,##0.00 €"
    cell_tot.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    cell_tot.fill = PatternFill("solid", fgColor=COULEUR_TOTAL_GEN)
    cell_tot.alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=COULEUR_TOTAL_GEN)
    for col in (1, 2, 3):
        ws.cell(row=row, column=col).border = BORDER
    ws.row_dimensions[row].height = 30
    row += 2

    # TVA et TTC
    tva = round(dpgf.montant_dqe_he * 0.20, 2)
    ttc = round(dpgf.montant_dqe_he + tva, 2)
    ws.cell(row=row, column=2, value="TVA 20 %")
    ws.cell(row=row, column=2).font = Font(name="Calibri", size=10, italic=True)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    cell_tva = ws.cell(row=row, column=3, value=tva)
    cell_tva.number_format = "#,##0.00 €"
    cell_tva.font = Font(name="Calibri", size=10, italic=True)
    cell_tva.alignment = Alignment(horizontal="right")
    row += 1

    ws.cell(row=row, column=2, value="TOTAL DQE TTC")
    ws.cell(row=row, column=2).font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    cell_ttc = ws.cell(row=row, column=3, value=ttc)
    cell_ttc.number_format = "#,##0.00 €"
    cell_ttc.font = Font(name="Calibri", size=11, bold=True)
    cell_ttc.alignment = Alignment(horizontal="right")


def _sheet_bpu(wb: Workbook, dpgf: DPGFGeneree, marche_ref: str, candidat: str) -> None:
    ws = wb.create_sheet("BPU")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18

    ws["A1"] = "Bordereau des Prix Unitaires"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=COULEUR_HEADER)
    ws.merge_cells("A1:D1")

    if marche_ref:
        ws["A2"] = f"Marché : {marche_ref}"
        ws["A2"].font = Font(name="Calibri", size=10, italic=True)
        ws.merge_cells("A2:D2")

    if candidat:
        ws["A3"] = f"Candidat : {candidat}"
        ws["A3"].font = Font(name="Calibri", size=10, italic=True)
        ws.merge_cells("A3:D3")

    ws["A4"] = f"Date : {date.today().strftime('%d/%m/%Y')}"
    ws["A4"].font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws.merge_cells("A4:D4")

    row = 6
    headers = ["Code", "Désignation", "Unité", "Prix unitaire HT (€)"]
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COULEUR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 24

    row += 1
    by_cat: dict[str, list] = defaultdict(list)
    for p in dpgf.bpu:
        by_cat[p.categorie].append(p)

    for cat, prestations in by_cat.items():
        c = ws.cell(row=row, column=1, value=cat)
        c.font = Font(name="Calibri", size=11, bold=True, color=COULEUR_HEADER)
        c.fill = PatternFill("solid", fgColor=COULEUR_CATEGORIE)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=0)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 20
        row += 1

        for p in prestations:
            ws.cell(row=row, column=1, value=p.code).alignment = Alignment(
                horizontal="center", vertical="top"
            )
            ws.cell(row=row, column=2, value=p.libelle).alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            ws.cell(row=row, column=3, value=p.unite).alignment = Alignment(
                horizontal="center", vertical="top"
            )
            cell_pu = ws.cell(row=row, column=4, value=p.prix_unitaire)
            cell_pu.number_format = "#,##0.00 €"
            cell_pu.alignment = Alignment(horizontal="right", vertical="top")
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = BORDER
                ws.cell(row=row, column=col).font = Font(name="Calibri", size=10)
            row += 1

    ws.freeze_panes = "A7"


def _sheet_dqe(wb: Workbook, dpgf: DPGFGeneree, marche_ref: str, candidat: str) -> None:
    ws = wb.create_sheet("DQE")

    widths = [8, 50, 14, 14, 14, 14]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws["A1"] = "Détail Quantitatif Estimatif"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=COULEUR_HEADER)
    ws.merge_cells("A1:F1")

    if marche_ref:
        ws["A2"] = f"Marché : {marche_ref}"
        ws["A2"].font = Font(name="Calibri", size=10, italic=True)
        ws.merge_cells("A2:F2")
    if candidat:
        ws["A3"] = f"Candidat : {candidat}"
        ws["A3"].font = Font(name="Calibri", size=10, italic=True)
        ws.merge_cells("A3:F3")
    ws["A4"] = (
        f"Programme indicatif : {dpgf.description_programme}"
        if dpgf.description_programme
        else ""
    )
    ws["A4"].font = Font(name="Calibri", size=10, italic=True, color="444444")
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A4:F4")
    ws.row_dimensions[4].height = 32

    row = 6
    headers = ["Code", "Désignation", "Unité", "Qté", "Prix unitaire HT (€)", "Montant HT (€)"]
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COULEUR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 28

    row += 1
    by_cat: dict[str, list] = defaultdict(list)
    for ligne in dpgf.dqe:
        by_cat[ligne.categorie].append(ligne)

    for cat, lignes in by_cat.items():
        c = ws.cell(row=row, column=1, value=cat)
        c.font = Font(name="Calibri", size=11, bold=True, color=COULEUR_HEADER)
        c.fill = PatternFill("solid", fgColor=COULEUR_CATEGORIE)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 20
        row += 1

        sous_total_cat = 0.0
        for ligne in lignes:
            est_pourcent_travaux = ligne.unite.strip().lower() in {"% travaux", "%travaux", "%"}
            ws.cell(row=row, column=1, value=ligne.code).alignment = Alignment(
                horizontal="center", vertical="top"
            )
            ws.cell(row=row, column=2, value=ligne.libelle).alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            ws.cell(row=row, column=3, value=ligne.unite).alignment = Alignment(
                horizontal="center", vertical="top"
            )
            cell_q = ws.cell(row=row, column=4, value=ligne.quantite)
            if est_pourcent_travaux:
                cell_q.number_format = "#,##0 €"  # quantité = montant travaux
            else:
                cell_q.number_format = "#,##0.00"
            cell_q.alignment = Alignment(horizontal="right", vertical="top")
            cell_pu = ws.cell(row=row, column=5, value=ligne.prix_unitaire)
            if est_pourcent_travaux:
                cell_pu.number_format = "0.00\\%"  # taux %
            else:
                cell_pu.number_format = "#,##0.00 €"
            cell_pu.alignment = Alignment(horizontal="right", vertical="top")
            cell_mt = ws.cell(row=row, column=6, value=ligne.montant)
            cell_mt.number_format = "#,##0.00 €"
            cell_mt.alignment = Alignment(horizontal="right", vertical="top")
            cell_mt.font = Font(name="Calibri", size=10, bold=True)

            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = BORDER
                if not cell.font.bold:
                    cell.font = Font(name="Calibri", size=10)

            sous_total_cat += ligne.montant
            row += 1

        c = ws.cell(row=row, column=2, value=f"Sous-total {cat}")
        c.font = Font(name="Calibri", size=10, bold=True, italic=True)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.fill = PatternFill("solid", fgColor=COULEUR_TOTAL_CAT)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell_tot = ws.cell(row=row, column=6, value=round(sous_total_cat, 2))
        cell_tot.number_format = "#,##0.00 €"
        cell_tot.font = Font(name="Calibri", size=10, bold=True)
        cell_tot.fill = PatternFill("solid", fgColor=COULEUR_TOTAL_CAT)
        cell_tot.alignment = Alignment(horizontal="right", vertical="center")
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    row += 1
    c = ws.cell(row=row, column=2, value="TOTAL DQE HT")
    c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COULEUR_TOTAL_GEN)
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell_tot = ws.cell(row=row, column=6, value=dpgf.montant_dqe_he)
    cell_tot.number_format = "#,##0.00 €"
    cell_tot.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    cell_tot.fill = PatternFill("solid", fgColor=COULEUR_TOTAL_GEN)
    cell_tot.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 26
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = BORDER

    ws.freeze_panes = "A7"
